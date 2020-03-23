import openpyxl
import os
from collections import defaultdict


#해당 위치의 파일리스트 호출
def get_file_list(path, file_format):
    file_list =[]
    if file_format == '.js':
        for root, dirs, files in os.walk(path):
            for file in files:
                if file.endswith(file_format) and root.startswith(path+'\\src' ):
                    folder = root.split('\\')[-2:]
                    file_list.append([root,folder[0]+'\\'+folder[1],file])
                else: pass
    else:
        for root, dirs, files in os.walk(path):
            for file in files:
                if file.endswith(file_format):
                    folder = root.split('\\')[-2:]
                    file_list.append([root,folder[0]+'\\'+folder[1],file])
                else: pass
    return file_list

#문자 판별, 기호만 있는 경우 제외
# "{data}" 형식의 요소 제외
def is_text(input_s):
    if input_s.startswith("@string"): #String 지정된 값 pass
        return False
    elif input_s.startswith("{") and input_s.endswith("}"):
        return False
    
    for c in input_s:
        if ord('가') <= ord(c) <= ord('힣'):
            return True
        elif ord('a') <= ord(c.lower()) <= ord('z'):
            return True
    return False

#get id, properties from files 
def get_id_properties(file_type,file_list):
    prop_list=[]
    i = 1 #index
    
    for file in file_list:
        file_path = file[0] + '\\'
        file_name = file[2]
        
        code = open(file_path+file_name,'rt', encoding='UTF8')
        while True:
            line = code.readline()
            if not line: break
            line = line.strip()
            if line.startswith("id="):
                object_id=line
            if line.startswith("text=") or line.startswith("hint=") or line.startswith("title=") or line.startswith("message="):
                object_id=object_id.replace("\"","").replace('id=','')
                text_type=line.split('=')[0]
                string = line.split("\"")[1]
                
                if is_text(string):
                    prop_list.append([i,file_type,"",file[1],file_name,object_id,text_type,string,"번역필요"+str(i)])
                    i+=1
                
                
        code.close()
        
    return prop_list


# 추출한 string excel 저장
def write_xl(filename, string_list):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1,column=1).value="Index"
    sheet.cell(row=1,column=2).value="Type"
    sheet.cell(row=1,column=3).value="StringID"
    sheet.cell(row=1,column=4).value="Path"
    sheet.cell(row=1,column=5).value="File"
    sheet.cell(row=1,column=6).value="ObjectID // Function"
    sheet.cell(row=1,column=7).value="TextType"
    sheet.cell(row=1,column=8).value="TextValue0"
    sheet.cell(row=1,column=9).value="TextValue1"
    
    sheet.column_dimensions['B'].width= 30
    sheet.column_dimensions['C'].width= 30
    sheet.column_dimensions['D'].width= 40
    sheet.column_dimensions['E'].width= 50
    sheet.column_dimensions['f'].width= 50
    sheet.column_dimensions['g'].width= 20
    sheet.column_dimensions['h'].width= 70
    sheet.column_dimensions['i'].width= 70
    
    for i in range(len(string_list)):
        sheet.cell(i+2,1).value=string_list[i][0]
        sheet.cell(i+2,2).value=string_list[i][1]
        sheet.cell(i+2,3).value=string_list[i][2]
        sheet.cell(i+2,4).value=string_list[i][3]
        sheet.cell(i+2,5).value=string_list[i][4]
        sheet.cell(i+2,6).value=string_list[i][5]
        sheet.cell(i+2,7).value=string_list[i][6]
        sheet.cell(i+2,8).value=string_list[i][7]
        sheet.cell(i+2,9).value=string_list[i][8]
    
    wb.save(filename+".xlsx")
    wb.close()    
    
    
## JS Logic settext parshing
def get_js_func_properties(file_list):
    prop_list=[]
    i=1

    for file in file_list:
        file_path = file[0] + '\\'
        file_name = file[2]

        code = open(file_path+file_name,'rt', encoding='UTF8')
        while True:
            line = code.readline()
            if not line: break
            line = line.strip()
            
            if 'Top.Dom.selectById("' in line:
                object_id = line.split('Top.Dom.selectById("')[1]

            if '.setText' in line:
                string = line.split('.setText')[-1]
                object_id=object_id.replace("\"","") 
                
                if '\"' in string :
                    string_ = string.split("\"")
                    for j in range(len(string_)):
                        if j%2 == 1 and is_text(string_[j]):
                            prop_list.append([i,'js','',file[1],file_name,object_id,'setText',string_[j]])
                            i+=1

        code.close()

    return prop_list



#모든 객체 string 값 출력
def extract_main(path,excel_file_name):
    path = path.replace('\\','\\')
    
    #1. Runtime Files
    path2 = path +'\\meta\\runtime'
    runtime_files = get_file_list(path2, '.xml')
    runtime_text = get_id_properties('runtime',runtime_files)
    
    #2 CustomWidget FILES
    path3 = path +'\\meta'
    custom_files = get_file_list(path3, 'tcw')
    custom_text = get_id_properties('custom',custom_files)
    
    # 3.menu files
    menu_files= get_file_list(path, 'tmrf')
    menu_text = get_id_properties('menu',menu_files)
    
    # 4.layout files
    layout_files= get_file_list(path, 'tlf')
    layout_text = get_id_properties('layout',layout_files)
    
    # 5. js files
    ## 보류 .... 
    js_text = []
    
#     js_list=get_file_list(path, '.js')
#     if js_list != [] :
#         js_text = get_js_func_properties(js_list)
#     else:
#         js_text = []
    
    all_list = menu_text+custom_text+runtime_text+layout_text+js_text
    
    write_xl(excel_file_name, all_list)
    

#string Object 출력

# 프로젝트 하위 String 파일 목록 추출
def get_string_file_list(path):
    file_list =[]
    path = path +'\\res'
    for root, dirs, files in os.walk(path):
        for file in files:
            if file.startswith('str'):
                file_list.append(root+'\\'+file)
    return file_list

#String id 및 value 읽어오기
def read_string_file(file_path):
    string_list = []
    code = open(file_path, 'rt', encoding='UTF8')
    while True:
        line = code.readline()
        if not line: break
        line = line.strip()
        if line.startswith("name="):
            str_id = line.split("\"")[1]
            str_val = line.split(str_id+'\">')[1].split('</resource:String>')[0]
            string_list.append([str_id,str_val])
        else: pass
    code.close()
        
    return string_list


#영문, 한글 두 str파일을 파싱하여 하나의 dict로 생성 
# key: id , value: string value list

def extract_match_string(files):
    # 파일별 string 추출
    file_dict = defaultdict(list)
    for file in files:
        file_dict[file] = read_string_file(file)
   

    string_dict = defaultdict(list)
    for file in file_dict :
        for string in file_dict[file]:
            str_id = string[0]
            str_val = string[1]
            if string_dict[str_id]:
                string_dict[str_id].append(str_val)
            else :
                string_dict[str_id]=[str_val]
    
    return string_dict


#string object 하나의 excel로 출력
def write_string_xl(filename, string_dict):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1,column=1).value="Index"
    sheet.cell(row=1,column=2).value="StringID"
    sheet.cell(row=1,column=3).value="en"
    sheet.cell(row=1,column=4).value="kr"
    
    sheet.column_dimensions['B'].width= 30
    sheet.column_dimensions['C'].width= 30
    sheet.column_dimensions['D'].width= 40
    
    n=1
    for str_id in string_dict:
        sheet.cell(n+1,1).value=n
        sheet.cell(n+1,2).value=str_id
        sheet.cell(n+1,3).value=string_dict[str_id][0]
        sheet.cell(n+1,4).value=string_dict[str_id][1]
        n+=1
    
    wb.save(filename+".xlsx")
    wb.close()  
    

#string dict에 string 존재 여부 확인
def isExist(string,string_dict):
    for item in string_dict:
        if string in string_dict[item]:
            return True
    return False



# string object에 출력한 excel-번역본 업데이트 적용 
def update_string_object(excel_file_name,path,string_object_name):
    
    string_object_path=path+'\\res'
    string_file_list= get_file_list(string_object_path, string_object_name+'.xml')
    files = []
    for file in string_file_list:
        files.append(file[0]+'//'+file[2])
        
    str_dict = extract_match_string(files)
    
    raw_string_path = './'+excel_file_name+'.xlsx'
    print(raw_string_path)
    raw_string_wb = openpyxl.load_workbook(r'./'+excel_file_name+'.xlsx', data_only=True)
    raw_string_ws = raw_string_wb['Sheet']
    
    print("file read")
    
    #raw_string save as list
    raw_string_data =[]
    for row in raw_string_ws:
        row_data =[]
        for cell in row:
            row_data.append(cell.value)
        raw_string_data.append(row_data)
        
    print("create string list")
    print("------------------------------------------")
    
    #raw string 추가
    for j in range(1,len(raw_string_data)):
        raw_string_value = raw_string_data[j][7]
        exist = isExist(raw_string_value,str_dict)
        if exist==False :
            # str_dict에 value 추가 & ID 부여
            index=len(str_dict)
            newid='String'+("%06d"% index) #id rule... need update....
            str_dict[newid]=[raw_string_data[j][7],raw_string_data[j][8]]
            print(newid,raw_string_data[j][7],raw_string_data[j][8])
            
    
    print("------------------------------------------")
    
    for i in range(2):
        code = open(string_file_list[i][0]+'\\'+string_file_list[i][2],'r+', encoding='UTF8')
        lines=code.readlines()
        code.close()
        lines = lines[:4]
        for item in str_dict:
            lines.append('        <resource:String\n')
            stridval='        name="'+item+'">'+str_dict[item][i]+'</resource:String>\n'
            lines.append(stridval)
        lines.append('</resource:Strings>')
        code = open(string_file_list[i][0]+'\\'+string_file_list[i][2],'w+t', encoding='UTF8')
        code.writelines(lines)
        code.close()
        print(string_file_list[i][0]+'\\'+string_file_list[i][2])
        
        
        
        
#모든 obj의 string value 파악 extract_main 변형
#개별 파일 단위로 변환
def change_string_val_to_id(file,str_dict,string_object):
    code = open(file,'r+',encoding='UTF8')
    lines = code.readlines()
    code.close()
    
    newlines=[]
    for line in lines:
        if 'text=\"' in line or 'hint=\"' in line or 'message=\"' in line or 'title=\"' in line :
            type_val=line.split('\"')[0]
            text_val=line.split('\"')[1]
            for key, value in str_dict.items():
                if text_val == value[0] and not text_val.startswith("@string/"):
                    print(line)
                    line = line.replace(type_val+'\"'+text_val+'\"' , type_val+"\"@string/"+string_object+"/"+key+'\"')
                    print(line)
        newlines.append(line)
    code = open(file,'w+t', encoding='UTF8')
    code.writelines(newlines)
    code.close()
    

#프로젝트 내의 모든 하드코딩 스트링을 id로 변환
def switch_all_file_string_to_id(path,string_object_name):
    #String object 읽어오기
    string_object_path=path+'\\res'
    string_file_list= get_file_list(string_object_path, string_object_name+'.xml')
    files = []
    for file in string_file_list:
        files.append(file[0]+'//'+file[2])
        
    str_dict = extract_match_string(files)
    
    #모든 obj 파일 리스트
    #js 파일 제외
    
    #1. Runtime Files
    path2 = path +'\\meta\\runtime'
    runtime_files = get_file_list(path2, '.xml')
    #2 CustomWidget FILES
    path3 = path +'\\meta'
    custom_files = get_file_list(path3, 'tcw')
    # 3.menu files
    menu_files= get_file_list(path, 'tmrf')
    # 4.layout files
    layout_files= get_file_list(path, 'tlf')
    
    object_files = runtime_files+custom_files+menu_files+layout_files
    
    for file in object_files:
        change_string_val_to_id(file[0]+'\\'+file[2], str_dict, string_object_name)